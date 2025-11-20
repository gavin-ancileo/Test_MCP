"""
Google Drive MCP Server Wrapper
Custom MCP server wrapper for Google Drive
This implements a simple MCP server interface that wraps Google Drive API calls
"""

import httpx
import json
import os
import io
from typing import Dict, Any, Optional, AsyncGenerator
from contextlib import asynccontextmanager

# Binary file text extraction
try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from PyPDF2 import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from mcp import ClientSession
    try:
        from mcp.types import Tool, TextContent
    except ImportError:
        # Fallback if types not available
        from typing import TypedDict
        class TextContent(TypedDict):
            type: str
            text: str
        class Tool(TypedDict):
            name: str
            description: str
            inputSchema: dict
    MCP_AVAILABLE = True
except ImportError:
    MCP_AVAILABLE = False


def extract_text_from_docx(binary_content: bytes) -> str:
    """Extract text from .docx file"""
    if not DOCX_AVAILABLE:
        return "[python-docx library not available - cannot extract text from .docx]"

    try:
        doc = Document(io.BytesIO(binary_content))
        text_parts = []
        for para in doc.paragraphs:
            if para.text.strip():
                text_parts.append(para.text)
        return "\n".join(text_parts)
    except Exception as e:
        return f"[Error extracting text from .docx: {str(e)}]"


def extract_text_from_pdf(binary_content: bytes) -> str:
    """Extract text from PDF file"""
    if not PDF_AVAILABLE:
        return "[PyPDF2 library not available - cannot extract text from PDF]"

    try:
        pdf_reader = PdfReader(io.BytesIO(binary_content))
        text_parts = []
        for page in pdf_reader.pages:
            text = page.extract_text()
            if text.strip():
                text_parts.append(text)
        return "\n".join(text_parts)
    except Exception as e:
        return f"[Error extracting text from PDF: {str(e)}]"


def extract_text_from_xlsx(binary_content: bytes) -> str:
    """Extract text from .xlsx file"""
    if not XLSX_AVAILABLE:
        return "[openpyxl library not available - cannot extract text from .xlsx]"

    try:
        wb = load_workbook(io.BytesIO(binary_content), read_only=True, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")

            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(row_values):  # Skip empty rows
                    text_parts.append("\t".join(row_values))

        return "\n".join(text_parts)
    except Exception as e:
        return f"[Error extracting text from .xlsx: {str(e)}]"


class DriveMCPWrapper:
    """
    Simple MCP server wrapper for Google Drive API
    Implements MCP protocol to wrap Google Drive REST API calls
    """

    def __init__(self, user_id: str, access_token: str):
        self.user_id = user_id
        self.access_token = access_token
        self.base_url = "https://www.googleapis.com/drive/v3"
        self.headers = {
            "Authorization": f"Bearer {access_token}",
            "Accept": "application/json"
        }

    def _raise_drive_error(self, response: httpx.Response, operation: str):
        """Raise descriptive error based on status code"""
        status_code = response.status_code

        if status_code == 401:
            raise Exception(
                f"Google Drive authentication failed (401). Your access token has expired. "
                f"Please reconnect Google Drive in Settings > Integrations to refresh the connection."
            )
        elif status_code == 403:
            raise Exception(
                f"Google Drive permission denied (403). You don't have permission to access this resource. "
                f"Please check file sharing settings or reconnect with appropriate permissions."
            )
        elif status_code == 404:
            raise Exception(
                f"Google Drive file not found (404).\n\n"
                f"Possible reasons:\n"
                f"1. File has been deleted or moved to trash\n"
                f"2. File ID is invalid or malformed\n"
                f"3. You don't have permission to access this file\n"
                f"4. File is in a shared drive you're not a member of\n"
                f"5. File was shared with you but sharing was revoked\n\n"
                f"Solutions:\n"
                f"- Use drive_list_files or drive_search_files to find accessible files\n"
                f"- Check if file is in trash and restore it\n"
                f"- Ask file owner to reshare the file with you"
            )
        elif status_code == 429:
            raise Exception(
                f"Google Drive rate limit exceeded (429). Too many requests in a short time. "
                f"Please wait a moment and try again."
            )
        else:
            # Generic error with full details
            error_detail = response.text
            try:
                error_json = response.json()
                error_message = error_json.get('error', {}).get('message', error_detail)
                raise Exception(f"Google Drive API error ({status_code}): {error_message}")
            except:
                raise Exception(f"Google Drive API error ({status_code}): {error_detail}")
    
    async def list_tools(self) -> list[dict]:
        """List available Google Drive MCP tools"""
        return [
            {
                "name": "drive_list_files",
                "description": "List files in Google Drive. Returns the list of files immediately. Do not say you will list files, just return the results.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string", "description": "Optional search query"},
                        "page_size": {"type": "integer", "description": "Number of results", "default": 10}
                    },
                    "required": []
                }
            },
            {
                "name": "drive_search_files",
                "description": "Search for files in Google Drive. Returns matching files immediately. Do not say you will search, just return the search results.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string", "description": "Search query"}
                    },
                    "required": ["query"]
                }
            },
            {
                "name": "drive_read_file",
                "description": "Read and return file contents from Google Drive. Returns the file content immediately. Do not say you will read the file, just return the content directly.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "file_id": {"type": "string", "description": "Google Drive file ID"}
                    },
                    "required": ["file_id"]
                }
            },
            {
                "name": "drive_create_file",
                "description": "Create a new file in Google Drive. Returns the created file details immediately. Do not say you will create a file, just return the creation result.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "File name"},
                        "mime_type": {"type": "string", "description": "MIME type", "default": "text/plain"},
                        "content": {"type": "string", "description": "File content"}
                    },
                    "required": ["name"]
                }
            },
            {
                "name": "drive_update_file",
                "description": "Update a file in Google Drive. Returns the updated file details immediately. Do not say you will update the file, just return the update result.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "file_id": {"type": "string", "description": "Google Drive file ID"},
                        "content": {"type": "string", "description": "New file content"}
                    },
                    "required": ["file_id", "content"]
                }
            },
            {
                "name": "drive_export_file",
                "description": "Export Google Docs/Sheets/Slides to PDF, DOCX, XLSX, etc. and read the content. Use this to scan/read PDF, DOC, or Sheet files. Returns exported content immediately.",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "file_id": {"type": "string", "description": "Google Drive file ID"},
                        "mime_type": {
                            "type": "string",
                            "description": "Export MIME type: application/pdf, application/vnd.openxmlformats-officedocument.wordprocessingml.document, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, text/plain, etc.",
                            "default": "application/pdf"
                        }
                    },
                    "required": ["file_id"]
                }
            },
            {
                "name": "drive_api_call",
                "description": "⭐ RECOMMENDED: Universal Google Drive REST API v3 caller - call ANY Drive API endpoint directly. Use this for advanced operations. Examples: GET /files/{fileId}?fields=* (detailed metadata), GET /files/{fileId}/permissions (sharing settings), POST /files/{fileId}/copy (copy file), etc. See Drive API docs: https://developers.google.com/drive/api/v3/reference",
                "inputSchema": {
                    "type": "object",
                    "properties": {
                        "endpoint": {
                            "type": "string",
                            "description": "API endpoint path without base URL. Examples: '/files/{fileId}', '/files/{fileId}/export', '/files/{fileId}/permissions'. Do NOT include 'https://www.googleapis.com/drive/v3' prefix."
                        },
                        "method": {
                            "type": "string",
                            "enum": ["GET", "POST", "PUT", "DELETE", "PATCH"],
                            "description": "HTTP method",
                            "default": "GET"
                        },
                        "params": {
                            "type": "object",
                            "description": "Query parameters (e.g., {\"fields\": \"*\", \"pageSize\": 100})",
                            "default": {}
                        },
                        "body": {
                            "type": "object",
                            "description": "Request body for POST/PUT/PATCH requests",
                            "default": {}
                        }
                    },
                    "required": ["endpoint"]
                }
            },
        ]

    async def call_tool(self, tool_name: str, arguments: Dict[str, Any]) -> Dict[str, Any]:
        """Call a Google Drive tool"""
        async with httpx.AsyncClient(timeout=30.0) as client:
            if tool_name == "drive_list_files":
                query = arguments.get("query")
                page_size = arguments.get("page_size", 10)
                
                params = {
                    "pageSize": page_size,
                    "fields": "files(id,name,mimeType,modifiedTime,size)",
                    "supportsAllDrives": "true",
                    "includeItemsFromAllDrives": "true"
                }
                if query:
                    params["q"] = f"name contains '{query}'"
                
                response = await client.get(
                    f"{self.base_url}/files",
                    headers=self.headers,
                    params=params
                )
                if response.status_code == 200:
                    data = response.json()
                    return {
                        "content": [
                            {
                                "type": "text",
                                "text": json.dumps({
                                    "files": [
                                        {
                                            "id": f.get("id"),
                                            "name": f.get("name"),
                                            "mimeType": f.get("mimeType", "unknown"),
                                            "modifiedTime": f.get("modifiedTime", ""),
                                            "size": f.get("size", "0")
                                        }
                                        for f in data.get("files", [])
                                    ]
                                })
                            }
                        ]
                    }
                else:
                    self._raise_drive_error(response, "drive_list_files")
            
            elif tool_name == "drive_search_files":
                query = arguments.get("query")
                params = {
                    "q": f"name contains '{query}' or fullText contains '{query}'",
                    "pageSize": 50,
                    "fields": "files(id,name,mimeType,modifiedTime)",
                    "supportsAllDrives": "true",
                    "includeItemsFromAllDrives": "true"
                }
                
                response = await client.get(
                    f"{self.base_url}/files",
                    headers=self.headers,
                    params=params
                )
                if response.status_code == 200:
                    data = response.json()
                    return {
                        "content": [
                            {
                                "type": "text",
                                "text": json.dumps({
                                    "files": [
                                        {
                                            "id": f.get("id"),
                                            "name": f.get("name"),
                                            "mimeType": f.get("mimeType", "unknown"),
                                            "modifiedTime": f.get("modifiedTime", "")
                                        }
                                        for f in data.get("files", [])
                                    ]
                                })
                            }
                        ]
                    }
                else:
                    self._raise_drive_error(response, "drive_search_files")

            elif tool_name == "drive_read_file":
                file_id = arguments.get("file_id")

                # Validate file_id before API call
                if not file_id or not isinstance(file_id, str) or len(file_id) < 10:
                    raise ValueError(
                        f"Invalid file_id: '{file_id}'. "
                        f"Expected a valid Google Drive file ID (e.g., '1ABcd...xyz'). "
                        f"Use drive_list_files or drive_search_files to get valid file IDs."
                    )

                # First get file metadata
                metadata_response = await client.get(
                    f"{self.base_url}/files/{file_id}",
                    headers=self.headers,
                    params={
                        "fields": "id,name,mimeType",
                        "supportsAllDrives": "true"
                    }
                )
                
                if metadata_response.status_code != 200:
                    self._raise_drive_error(metadata_response, "drive_read_file (metadata)")
                
                metadata = metadata_response.json()
                mime_type = metadata.get("mimeType", "")
                
                # Download file content
                # For Google Workspace files, use export
                if mime_type.startswith("application/vnd.google-apps"):
                    # Determine best export format based on file type
                    if "document" in mime_type:
                        export_mime = "text/plain"
                    elif "spreadsheet" in mime_type:
                        export_mime = "text/csv"
                    elif "presentation" in mime_type:
                        export_mime = "application/pdf"  # Slides export to PDF
                    elif "drawing" in mime_type:
                        export_mime = "application/pdf"  # Drawings export to PDF
                    elif "form" in mime_type:
                        export_mime = "application/zip"  # Forms export to ZIP
                    else:
                        export_mime = "text/plain"  # Default fallback

                    content_response = await client.get(
                        f"{self.base_url}/files/{file_id}/export",
                        headers=self.headers,
                        params={
                            "mimeType": export_mime,
                            "supportsAllDrives": "true"
                        }
                    )
                else:
                    content_response = await client.get(
                        f"{self.base_url}/files/{file_id}",
                        headers={**self.headers, "Accept": "*/*"},
                        params={
                            "alt": "media",
                            "supportsAllDrives": "true"
                        }
                    )
                
                if content_response.status_code == 200:
                    # Determine if content is binary and needs text extraction
                    file_name = metadata.get("name", "").lower()
                    binary_content = content_response.content

                    # Check for unsupported .doc files (old Word format)
                    if file_name.endswith('.doc') and not file_name.endswith('.docx'):
                        return json.dumps({
                            "error": "unsupported_format",
                            "message": f"⚠️ Old Word format (.doc) is not supported.\n\n"
                                       f"The file '{metadata.get('name')}' uses the legacy .doc format which cannot be read directly.\n\n"
                                       f"Please either:\n"
                                       f"1. Convert to .docx format (File → Save As → .docx)\n"
                                       f"2. Open in Google Docs (will auto-convert)\n"
                                       f"3. Save as PDF for reading\n\n"
                                       f"After converting, ask me to read the file again.",
                            "file_name": metadata.get("name"),
                            "file_id": file_id
                        })

                    # Extract text from binary files
                    if file_name.endswith('.docx') or 'wordprocessingml' in mime_type:
                        content = extract_text_from_docx(binary_content)
                    elif file_name.endswith('.pdf') or mime_type == 'application/pdf':
                        content = extract_text_from_pdf(binary_content)
                    elif file_name.endswith('.xlsx') or 'spreadsheetml' in mime_type or 'vnd.google-apps.spreadsheet' in mime_type:
                        # Handle both .xlsx files AND Google Sheets (exported as CSV/text)
                        # For Google Sheets, content is already CSV text, not binary Excel
                        if 'vnd.google-apps.spreadsheet' in mime_type:
                            # Already exported as CSV text, just use it
                            content = content_response.text if hasattr(content_response, 'text') else str(binary_content, 'utf-8')
                        else:
                            # Real .xlsx binary file, extract with openpyxl
                            content = extract_text_from_xlsx(binary_content)
                    else:
                        # For text files or other formats, use text content
                        content = content_response.text if hasattr(content_response, 'text') else str(binary_content)

                    # File size limits: 2MB max content, 3MB hard limit
                    MAX_CONTENT_SIZE = 2 * 1024 * 1024  # 2MB
                    HARD_LIMIT_SIZE = 3 * 1024 * 1024  # 3MB

                    # Check if file is too large (>3MB)
                    if len(content) > HARD_LIMIT_SIZE:
                        file_size_mb = len(content) / (1024 * 1024)
                        raise ValueError(
                            f"File too large ({file_size_mb:.1f}MB). "
                            f"Maximum file size is 3MB. "
                            f"Please use a smaller file or export to a different format."
                        )

                    return {
                        "content": [
                            {
                                "type": "text",
                                "text": json.dumps({
                                    "file_id": file_id,
                                    "name": metadata.get("name", ""),
                                    "mimeType": mime_type,
                                    "content": content[:MAX_CONTENT_SIZE],  # Limit content size to 2MB
                                    "truncated": len(content) > MAX_CONTENT_SIZE
                                })
                            }
                        ]
                    }
                else:
                    self._raise_drive_error(content_response, "drive_read_file (content)")

            elif tool_name == "drive_create_file":
                name = arguments.get("name")
                mime_type = arguments.get("mime_type", "text/plain")
                content = arguments.get("content", "")
                
                # Create file metadata
                metadata = {
                    "name": name,
                    "mimeType": mime_type
                }
                
                # Upload file
                files = {"metadata": (None, json.dumps(metadata), "application/json")}
                if content:
                    files["data"] = (name, content, mime_type)
                
                response = await client.post(
                    f"{self.base_url}/files",
                    headers=self.headers,
                    files=files if content else None,
                    data=json.dumps(metadata) if not content else None
                )
                
                if response.status_code in [200, 201]:
                    created = response.json()
                    return {
                        "content": [
                            {
                                "type": "text",
                                "text": json.dumps({
                                    "id": created.get("id"),
                                    "name": created.get("name"),
                                    "mimeType": created.get("mimeType")
                                })
                            }
                        ]
                    }
                else:
                    self._raise_drive_error(response, "drive_create_file")

            elif tool_name == "drive_update_file":
                file_id = arguments.get("file_id")
                content = arguments.get("content")
                
                # Update file content
                response = await client.patch(
                    f"{self.base_url}/files/{file_id}",
                    headers={**self.headers, "Content-Type": "text/plain"},
                    params={"uploadType": "media"},
                    content=content
                )
                
                if response.status_code == 200:
                    updated = response.json()
                    return {
                        "content": [
                            {
                                "type": "text",
                                "text": json.dumps({
                                    "id": updated.get("id"),
                                    "name": updated.get("name"),
                                    "success": True
                                })
                            }
                        ]
                    }
                else:
                    self._raise_drive_error(response, "drive_update_file")

            elif tool_name == "drive_export_file":
                file_id = arguments.get("file_id")
                mime_type = arguments.get("mime_type", "application/pdf")

                # Export file to specified format
                response = await client.get(
                    f"{self.base_url}/files/{file_id}/export",
                    headers=self.headers,
                    params={
                        "mimeType": mime_type,
                        "supportsAllDrives": "true"
                    }
                )

                if response.status_code == 200:
                    # For text-based exports (PDF as text, plain text, etc.)
                    content = response.text if hasattr(response, 'text') else str(response.content)

                    return {
                        "content": [
                            {
                                "type": "text",
                                "text": json.dumps({
                                    "file_id": file_id,
                                    "export_mime_type": mime_type,
                                    "content": content[:50000],  # Limit to 50KB for large files
                                    "content_size": len(content),
                                    "truncated": len(content) > 50000
                                })
                            }
                        ]
                    }
                else:
                    self._raise_drive_error(response, "drive_export_file")

            elif tool_name == "drive_api_call":
                # Generic API caller - allows calling any Drive REST API v3 endpoint
                endpoint = arguments.get("endpoint", "").lstrip("/")  # Remove leading slash
                method = arguments.get("method", "GET").upper()
                params = arguments.get("params", {})
                body = arguments.get("body", {})

                # Construct full URL
                url = f"{self.base_url}/{endpoint}"

                # Make the API call based on method
                if method == "GET":
                    response = await client.get(url, headers=self.headers, params=params)
                elif method == "POST":
                    response = await client.post(url, headers=self.headers, params=params, json=body if body else None)
                elif method == "PUT":
                    response = await client.put(url, headers=self.headers, params=params, json=body if body else None)
                elif method == "DELETE":
                    response = await client.delete(url, headers=self.headers, params=params)
                elif method == "PATCH":
                    response = await client.patch(url, headers=self.headers, params=params, json=body if body else None)
                else:
                    raise ValueError(f"Unsupported HTTP method: {method}")

                # Return response
                if response.status_code in [200, 201, 204]:
                    try:
                        data = response.json() if response.text else {}
                        return {
                            "content": [
                                {
                                    "type": "text",
                                    "text": json.dumps({
                                        "status_code": response.status_code,
                                        "data": data,
                                        "success": True
                                    })
                                }
                            ]
                        }
                    except json.JSONDecodeError:
                        # Return raw text if not JSON
                        return {
                            "content": [
                                {
                                    "type": "text",
                                    "text": json.dumps({
                                        "status_code": response.status_code,
                                        "data": response.text,
                                        "success": True
                                    })
                                }
                            ]
                        }
                else:
                    # Error response
                    self._raise_drive_error(response, f"drive_api_call ({method} {endpoint})")

            else:
                raise ValueError(f"Unknown Drive tool: {tool_name}")


@asynccontextmanager
async def get_drive_mcp_wrapper_session(user_id: str, access_token: str) -> AsyncGenerator[ClientSession, None]:
    """
    Get a Google Drive MCP wrapper session
    """
    if not MCP_AVAILABLE:
        raise RuntimeError("MCP library not available")
    
    wrapper = DriveMCPWrapper(user_id, access_token)
    
    # Create a simple session-like object
    class DriveMCPSession:
        def __init__(self, wrapper):
            self.wrapper = wrapper
        
        async def call_tool(self, name: str, arguments: Dict[str, Any]) -> Dict[str, Any]:
            return await self.wrapper.call_tool(name, arguments)
        
        async def list_tools(self):
            return await self.wrapper.list_tools()
    
    session = DriveMCPSession(wrapper)
    try:
        yield session
    finally:
        pass  # Cleanup if needed

